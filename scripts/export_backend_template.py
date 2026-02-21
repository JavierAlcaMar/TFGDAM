#!/usr/bin/env python3
"""
Exporta un modulo del backend SARA a la plantilla oficial de Excel.

Flujo:
1) Lee datos del backend (preview + reportes de evaluacion).
2) Rellena la plantilla vacia oficial.
3) Guarda un .xlsx de salida.

Requiere:
- Python 3.10+
- openpyxl (`pip install openpyxl`)
"""

from __future__ import annotations

import argparse
import json
import re
import urllib.error
import urllib.parse
import urllib.request
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Falta dependencia 'openpyxl'. Instala con: pip install openpyxl"
    ) from exc


SHEET_DATOS = "Datos Iniciales"
SHEET_ACTIVIDADES = "Actividades"
SHEET_EVALUACIONES = "Evaluaciones"

RA_ROW_START = 9
RA_MAX = 10
UT_ROW_START = 9
UT_MAX = 20
ACT_ROW_START = 35
ACT_MAX = 30
STUDENT_ROW_START_DATOS = 9
STUDENT_ROW_START_ACT = 6
STUDENT_ROW_START_EVAL = 5
STUDENT_MAX = 45  # plantilla oficial: alumnos 1..45

COL_RA_CODE = 1
COL_RA_NAME = 2
COL_RA_WEIGHT = 3
COL_UT_KEY = 5
COL_RA_START = 6  # F
COL_ACTIVITY_NAME = 5
COL_ACTIVITY_UT = 16  # P
COL_ACTIVITY_EVAL = 17  # Q
COL_STUDENT_CODE = 19  # S
COL_STUDENT_NAME = 20  # T

ACT_BLOCK_START_COL = 3  # C
ACT_BLOCK_WIDTH = 11
ACT_BLOCK_COUNT = 30
ACT_NAME_COL_OFFSET = 4  # C + 4 -> G (celda "Nombre" del bloque)

EVAL_COL_STUDENT_NAME = 2  # B
EVAL_COL_NUMERIC_BASE = 18  # R
EVAL_COL_SUGGESTED_BASE = 20  # T
EVAL_COL_FAILED_BASE = 31  # AE
EVAL_BLOCK_WIDTH = 15


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Rellena source_template.xlsx con datos del backend SARA."
    )
    parser.add_argument("--base-url", default="http://localhost:8080", help="Base URL del backend.")
    parser.add_argument("--token", required=True, help="JWT Bearer token.")
    parser.add_argument("--module-id", required=True, type=int, help="ID del modulo.")
    parser.add_argument(
        "--template",
        required=True,
        type=Path,
        help="Ruta del source_template.xlsx vacio.",
    )
    parser.add_argument(
        "--output",
        required=True,
        type=Path,
        help="Ruta de salida del .xlsx rellenado.",
    )
    parser.add_argument("--timeout", default=30, type=int, help="Timeout HTTP en segundos.")
    return parser.parse_args()


def natural_key(text: str) -> list[object]:
    parts = re.split(r"(\d+)", text or "")
    key: list[object] = []
    for part in parts:
        if part.isdigit():
            key.append(int(part))
        else:
            key.append(part.lower())
    return key


def as_decimal(value: object, default: Decimal = Decimal("0")) -> Decimal:
    if value is None:
        return default
    try:
        return Decimal(str(value))
    except Exception:
        return default


def scale(value: Decimal, places: int = 4) -> Decimal:
    quant = Decimal("1").scaleb(-places)
    return value.quantize(quant, rounding=ROUND_HALF_UP)


def to_percent_fraction(value: object) -> float:
    # Backend usa porcentaje 0..100, la plantilla usa celda porcentaje (0..1).
    raw = as_decimal(value)
    return float(scale(raw / Decimal("100"), 6))


def to_number(value: object, places: int = 4) -> float:
    return float(scale(as_decimal(value), places))


def api_get(base_url: str, path: str, token: str, timeout: int) -> dict:
    url = urllib.parse.urljoin(base_url.rstrip("/") + "/", path.lstrip("/"))
    request = urllib.request.Request(
        url=url,
        method="GET",
        headers={
            "Authorization": f"Bearer {token}",
            "Accept": "application/json",
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            payload = response.read().decode("utf-8")
            return json.loads(payload)
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {exc.code} en {path}: {detail}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"No se pudo conectar a {url}: {exc}") from exc


def validate_capacity(name: str, count: int, max_count: int) -> None:
    if count > max_count:
        raise RuntimeError(f"{name}: {count} supera el maximo soportado por la plantilla ({max_count}).")


def is_formula_cell(cell) -> bool:
    value = cell.value
    return isinstance(value, str) and value.startswith("=")


def set_if_writable(ws, row: int, column: int, value: object) -> None:
    cell = ws.cell(row=row, column=column)
    if isinstance(cell, MergedCell) or cell.__class__.__name__ == "MergedCell":
        return
    if is_formula_cell(cell):
        return
    try:
        cell.value = value
    except AttributeError:
        # Some template cells are merged/read-only placeholders.
        return


def optional_to_number(value: object, places: int = 2) -> float | None:
    if value is None:
        return None
    if isinstance(value, str) and not value.strip():
        return None
    return to_number(value, places)


def as_excel_student_code(value: object) -> object:
    if value is None:
        return None
    text = str(value).strip()
    if re.fullmatch(r"\d+", text):
        try:
            return int(text)
        except Exception:
            return text
    return text


def row_has_positive_ra_distribution(ws_datos, row: int) -> bool:
    for offset in range(RA_MAX):
        col = COL_RA_START + offset
        if as_decimal(ws_datos.cell(row=row, column=col).value) > 0:
            return True
    return False


def set_activity_grade(ws_actividades, row: int, start_col: int, grade_payload: dict) -> None:
    if not isinstance(grade_payload, dict):
        return

    exercise_values: dict[int, float] = {}
    for item in grade_payload.get("exerciseGrades") or []:
        if not isinstance(item, dict):
            continue
        idx_raw = item.get("exerciseIndex")
        try:
            idx = int(idx_raw)
        except Exception:
            continue
        if idx < 1 or idx > 10:
            continue
        value = optional_to_number(item.get("gradeValue"), 2)
        if value is None:
            continue
        exercise_values[idx] = value

    if exercise_values:
        for offset in range(1, ACT_BLOCK_WIDTH):
            set_if_writable(ws_actividades, row, start_col + offset, exercise_values.get(offset))
        main_grade = optional_to_number(grade_payload.get("gradeValue"), 2)
        if main_grade is not None:
            set_if_writable(ws_actividades, row, start_col, main_grade)
        return

    grade = optional_to_number(grade_payload.get("gradeValue"), 2)
    if grade is None:
        return

    exercise_cols = [start_col + offset for offset in range(1, ACT_BLOCK_WIDTH)]
    weighted_cols: list[int] = []
    for col in exercise_cols:
        weight = as_decimal(ws_actividades.cell(row=4, column=col).value)
        if weight > 0:
            weighted_cols.append(col)

    target_cols = weighted_cols if weighted_cols else exercise_cols

    # Keep formula-driven templates functional: main cell may be formula,
    # so we write exercises and let formula compute the instrument grade.
    for col in exercise_cols:
        set_if_writable(ws_actividades, row, col, grade if col in target_cols else None)
    set_if_writable(ws_actividades, row, start_col, grade)


def clear_ranges(ws_actividades, ws_evaluaciones, max_eval_period: int) -> None:
    # Actividades
    for row in range(STUDENT_ROW_START_ACT, STUDENT_ROW_START_ACT + STUDENT_MAX):
        set_if_writable(ws_actividades, row, 2, None)
        for block in range(ACT_BLOCK_COUNT):
            start_col = ACT_BLOCK_START_COL + (block * ACT_BLOCK_WIDTH)
            for offset in range(ACT_BLOCK_WIDTH):
                set_if_writable(ws_actividades, row, start_col + offset, None)

    # Evaluaciones
    clear_periods = max(1, max_eval_period)
    for row in range(STUDENT_ROW_START_EVAL, STUDENT_ROW_START_EVAL + STUDENT_MAX):
        set_if_writable(ws_evaluaciones, row, EVAL_COL_STUDENT_NAME, None)
        for period in range(1, clear_periods + 1):
            offset = (period - 1) * EVAL_BLOCK_WIDTH
            set_if_writable(ws_evaluaciones, row, EVAL_COL_NUMERIC_BASE + offset, None)
            set_if_writable(ws_evaluaciones, row, EVAL_COL_SUGGESTED_BASE + offset, None)
            set_if_writable(ws_evaluaciones, row, EVAL_COL_FAILED_BASE + offset, None)


def main() -> None:
    args = parse_args()

    if not args.template.exists():
        raise SystemExit(f"No existe plantilla: {args.template}")

    preview = api_get(args.base_url, f"/modules/{args.module_id}/preview", args.token, args.timeout)

    ras = sorted(preview.get("ras", []), key=lambda item: natural_key(str(item.get("code", ""))))
    uts = list(preview.get("uts", []))
    ut_ra_links = list(preview.get("utRaLinks", []))
    instruments = list(preview.get("instruments", []))
    students = sorted(preview.get("students", []), key=lambda item: natural_key(str(item.get("studentCode", ""))))
    grades = list(preview.get("grades", []))

    validate_capacity("RAs", len(ras), RA_MAX)
    validate_capacity("UTs", len(uts), UT_MAX)
    validate_capacity("Instrumentos", len(instruments), ACT_MAX)
    validate_capacity("Alumnos", len(students), STUDENT_MAX)

    periods = sorted(
        {int(item.get("evaluationPeriod")) for item in uts if item.get("evaluationPeriod") is not None}
    )
    evaluation_reports: dict[int, dict[int, dict]] = {}
    for period in periods:
        report = api_get(args.base_url, f"/modules/{args.module_id}/reports/evaluation/{period}", args.token, args.timeout)
        by_student: dict[int, dict] = {}
        for row in report.get("students", []):
            student_id = row.get("studentId")
            if student_id is not None:
                by_student[int(student_id)] = row
        evaluation_reports[period] = by_student

    ra_index_by_id = {int(ra["id"]): idx for idx, ra in enumerate(ras) if ra.get("id") is not None}
    ut_by_id = {int(ut["id"]): ut for ut in uts if ut.get("id") is not None}
    ut_order = {int(ut["id"]): idx for idx, ut in enumerate(uts) if ut.get("id") is not None}

    # Orden estable: UT -> activityId -> instrumentId
    instruments.sort(
        key=lambda ins: (
            ut_order.get(int(ins.get("utId", -1)), 9999),
            int(ins.get("activityId", 0)),
            int(ins.get("id", 0)),
        )
    )

    # Duplicados de nombre de instrumento rompen el mapeo de bloques en import.
    seen_instrument_names: set[str] = set()
    for ins in instruments:
        name = str(ins.get("name") or "").strip().lower()
        if name in seen_instrument_names:
            raise RuntimeError(f"Nombre de instrumento duplicado en modulo: '{ins.get('name')}'.")
        seen_instrument_names.add(name)

    grade_by_student_and_instrument: dict[tuple[int, int], dict] = {}
    for grade in grades:
        student_id = grade.get("studentId")
        instrument_id = grade.get("instrumentId")
        if student_id is None or instrument_id is None:
            continue
        grade_by_student_and_instrument[(int(student_id), int(instrument_id))] = grade

    wb = load_workbook(args.template)
    ws_datos = wb[SHEET_DATOS]
    ws_actividades = wb[SHEET_ACTIVIDADES]
    ws_evaluaciones = wb[SHEET_EVALUACIONES]

    clear_ranges(ws_actividades, ws_evaluaciones, max(periods) if periods else 1)

    # Cabecera minima
    set_if_writable(ws_datos, 3, 2, preview.get("moduleName"))

    # Tabla RA
    for idx, ra in enumerate(ras):
        row = RA_ROW_START + idx
        set_if_writable(ws_datos, row, COL_RA_CODE, ra.get("code"))
        set_if_writable(ws_datos, row, COL_RA_NAME, ra.get("name"))
        set_if_writable(ws_datos, row, COL_RA_WEIGHT, to_percent_fraction(ra.get("weightPercent")))

    # Tabla UT + reparto UT-RA
    links_by_ut: dict[int, list[dict]] = {}
    for link in ut_ra_links:
        ut_id = link.get("utId")
        if ut_id is None:
            continue
        links_by_ut.setdefault(int(ut_id), []).append(link)

    for idx, ut in enumerate(uts):
        ut_id = ut.get("id")
        if ut_id is None:
            continue
        row = UT_ROW_START + idx
        set_if_writable(ws_datos, row, COL_UT_KEY, ut.get("name"))
        for offset in range(RA_MAX):
            set_if_writable(ws_datos, row, COL_RA_START + offset, 0)
        for link in links_by_ut.get(int(ut_id), []):
            ra_id = link.get("raId")
            if ra_id is None:
                continue
            ra_idx = ra_index_by_id.get(int(ra_id))
            if ra_idx is None:
                continue
            col = COL_RA_START + ra_idx
            set_if_writable(ws_datos, row, col, to_percent_fraction(link.get("percent")))

    # Tabla actividades/instrumentos
    for idx, instrument in enumerate(instruments):
        row = ACT_ROW_START + idx
        ut_id = instrument.get("utId")
        ut = ut_by_id.get(int(ut_id)) if ut_id is not None else None

        set_if_writable(ws_datos, row, COL_ACTIVITY_NAME, instrument.get("name"))
        set_if_writable(ws_datos, row, COL_ACTIVITY_UT, ut.get("name") if ut else None)
        set_if_writable(ws_datos, row, COL_ACTIVITY_EVAL, ut.get("evaluationPeriod") if ut else None)

        linked_ra_ids = [
            int(ra_id)
            for ra_id in (instrument.get("raIds") or [])
            if ra_id is not None and int(ra_id) in ra_index_by_id
        ]
        if not linked_ra_ids:
            raise RuntimeError(f"Instrumento sin RA asociado: {instrument.get('name')}")

        preserve_existing_distribution = row_has_positive_ra_distribution(ws_datos, row)
        if not preserve_existing_distribution:
            for offset in range(RA_MAX):
                set_if_writable(ws_datos, row, COL_RA_START + offset, 0)

            instrument_weight_fraction = as_decimal(instrument.get("weightPercent")) / Decimal("100")
            per_ra = scale(instrument_weight_fraction / Decimal(len(linked_ra_ids)), 6)
            for ra_id in linked_ra_ids:
                col = COL_RA_START + ra_index_by_id[ra_id]
                set_if_writable(ws_datos, row, col, float(per_ra))

    # Estudiantes (Datos Iniciales)
    for idx, student in enumerate(students):
        row = STUDENT_ROW_START_DATOS + idx
        set_if_writable(ws_datos, row, COL_STUDENT_CODE, as_excel_student_code(student.get("studentCode")))
        set_if_writable(ws_datos, row, COL_STUDENT_NAME, student.get("fullName"))

    # Hoja Actividades: bloque por instrumento
    for block_idx, instrument in enumerate(instruments):
        start_col = ACT_BLOCK_START_COL + (block_idx * ACT_BLOCK_WIDTH)
        name_cell = ws_actividades.cell(row=1, column=start_col + ACT_NAME_COL_OFFSET)
        existing_name = name_cell.value
        if existing_name in (None, ""):
            set_if_writable(ws_actividades, 1, start_col + ACT_NAME_COL_OFFSET, instrument.get("name"))
        for weight_item in instrument.get("exerciseWeights") or []:
            if not isinstance(weight_item, dict):
                continue
            idx_raw = weight_item.get("exerciseIndex")
            try:
                exercise_idx = int(idx_raw)
            except Exception:
                continue
            if exercise_idx < 1 or exercise_idx > 10:
                continue
            weight = optional_to_number(weight_item.get("weightPercent"), 4)
            if weight is None:
                continue
            set_if_writable(ws_actividades, 4, start_col + exercise_idx, weight)

    for student_idx, student in enumerate(students):
        row = STUDENT_ROW_START_ACT + student_idx
        set_if_writable(ws_actividades, row, 2, student.get("fullName"))
        student_id = student.get("id")
        if student_id is None:
            continue
        for block_idx, instrument in enumerate(instruments):
            instrument_id = instrument.get("id")
            if instrument_id is None:
                continue
            grade_payload = grade_by_student_and_instrument.get((int(student_id), int(instrument_id)))
            if grade_payload is None:
                continue
            start_col = ACT_BLOCK_START_COL + (block_idx * ACT_BLOCK_WIDTH)
            set_activity_grade(ws_actividades, row, start_col, grade_payload)

    # Hoja Evaluaciones: valores backend (override o calculados segun API de reporte)
    for student_idx, student in enumerate(students):
        row = STUDENT_ROW_START_EVAL + student_idx
        set_if_writable(ws_evaluaciones, row, EVAL_COL_STUDENT_NAME, student.get("fullName"))
        student_id = student.get("id")
        if student_id is None:
            continue

        for period in periods:
            report_row = evaluation_reports.get(period, {}).get(int(student_id))
            if not report_row:
                continue
            offset = (period - 1) * EVAL_BLOCK_WIDTH
            set_if_writable(ws_evaluaciones, row, EVAL_COL_NUMERIC_BASE + offset, to_number(
                report_row.get("numericGrade"), 4
            ))
            set_if_writable(ws_evaluaciones, row, EVAL_COL_SUGGESTED_BASE + offset, int(
                report_row.get("suggestedBulletinGrade") or 1
            ))
            set_if_writable(ws_evaluaciones, row, EVAL_COL_FAILED_BASE + offset, (
                0 if bool(report_row.get("allRAsPassed")) else 1
            ))

    output_path = args.output.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"Plantilla exportada: {output_path}")
    print(
        f"Resumen -> RAs:{len(ras)} UTs:{len(uts)} instrumentos:{len(instruments)} alumnos:{len(students)} evaluaciones:{len(periods)}"
    )


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        raise SystemExit(130)
